"""Fichas de paciente: lo único que el agente sabe antes de marcar.

Supuesto 2 del plan (README §Supuestos): el agente llega a la llamada con lo que
tendría un sistema conectado al HIS — identificación, procedimiento, fecha de
cirugía, edad, comorbilidades, EPS — y **nada más**.

Este módulo carga a propósito solo dos de los cuatro xlsx del dataset:

    perfiles_pacientes_co.xlsx                      identificación y contacto
    perfiles_clinicos_pacientes_silver_contest.xlsx procedimiento y comorbilidades

Los otros dos (`trayectorias_postop_silver.xlsx`, con dolor/fiebre/herida reales,
y `dataset_final.xlsx`, con `label_ground_truth`) NO se cargan aquí ni son
importables desde `app/`: viven en `evals/dataset.py` y solo los usa el evaluador.
La separación es física, no una promesa: si el runtime no puede leer la
trayectoria, no puede filtrarse en una respuesta.
"""

from __future__ import annotations

import json
import warnings
from dataclasses import dataclass, field
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
from typing import Any

from app.config import get_settings

# openpyxl se queja de que los xlsx del kit no traen estilo por defecto. Es inocuo.
warnings.filterwarnings("ignore", message="Workbook contains no default style")

DIAS_POSTOP = (1, 3, 7, 14)


def leer_xlsx(ruta: Path) -> list[dict[str, Any]]:
    """Lee la primera hoja de un xlsx como lista de diccionarios."""
    from openpyxl import load_workbook

    libro = load_workbook(ruta, read_only=True, data_only=True)
    try:
        hoja = libro.worksheets[0]
        filas = hoja.iter_rows(values_only=True)
        cabecera = [str(c) if c is not None else "" for c in next(filas)]
        return [dict(zip(cabecera, fila)) for fila in filas if any(v is not None for v in fila)]
    finally:
        libro.close()


def _como_fecha(valor: Any) -> date | None:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str) and valor.strip():
        try:
            return datetime.fromisoformat(valor.strip()).date()
        except ValueError:
            return None
    return None


def _como_lista(valor: Any) -> list[str]:
    """`comorbilidades` viene como texto JSON: '["obesidad","hipertension"]'."""
    if isinstance(valor, list):
        return [str(v) for v in valor]
    if isinstance(valor, str) and valor.strip():
        try:
            cargado = json.loads(valor)
            return [str(v) for v in cargado] if isinstance(cargado, list) else [valor]
        except json.JSONDecodeError:
            return [valor]
    return []


@dataclass(frozen=True)
class Paciente:
    """Ficha administrativa y clínica de base. Sin estado postoperatorio."""

    paciente_id: str
    nombre_completo: str
    documento_cc: str
    eps: str
    ciudad: str
    departamento: str
    direccion: str
    edad: int
    genero: str
    procedimiento: str
    modulo_synthea: str
    fecha_cirugia: date | None
    comorbilidades: list[str] = field(default_factory=list)

    @property
    def nombre_pila(self) -> str:
        """«Mauricio Juan González Sánchez» → «Mauricio»: como lo llamaría una enfermera."""
        return self.nombre_completo.split()[0] if self.nombre_completo else ""

    @property
    def apellido(self) -> str:
        partes = self.nombre_completo.split()
        return partes[2] if len(partes) >= 4 else (partes[-1] if len(partes) > 1 else "")

    @property
    def saludo(self) -> str:
        """Cómo se dirige el agente al paciente.

        Sin «don/doña» a propósito. El campo `genero` del dataset no concuerda con
        el nombre en varios registros (p. ej. `pac_42_00000`, «Mauricio», marcado
        como F), así que un tratamiento derivado de ese campo produciría «doña
        Mauricio» en mitad de la llamada. Nombre y apellido con usted es natural en
        Colombia y no puede fallar por un dato malo.
        """
        return f"{self.nombre_pila} {self.apellido}".strip() or self.nombre_completo

    def fecha_de_llamada(self, dia_postop: int) -> date | None:
        """Fecha en la que correspondería la llamada del día postoperatorio `n`."""
        if not self.fecha_cirugia:
            return None
        return date.fromordinal(self.fecha_cirugia.toordinal() + dia_postop)

    def como_dict(self) -> dict[str, Any]:
        return {
            "paciente_id": self.paciente_id,
            "nombre_completo": self.nombre_completo,
            "documento_cc": self.documento_cc,
            "eps": self.eps,
            "ciudad": self.ciudad,
            "departamento": self.departamento,
            "edad": self.edad,
            "genero": self.genero,
            "procedimiento": self.procedimiento,
            "modulo_synthea": self.modulo_synthea,
            "fecha_cirugia": self.fecha_cirugia.isoformat() if self.fecha_cirugia else None,
            "comorbilidades": self.comorbilidades,
        }


@lru_cache
def cargar_pacientes() -> dict[str, Paciente]:
    """Los 40 pacientes del dataset, indexados por `paciente_id`."""
    s = get_settings()
    contacto = {f["paciente_id"]: f for f in leer_xlsx(s.dir_dataset / "perfiles_pacientes_co.xlsx")}
    clinicos = leer_xlsx(s.dir_dataset / "perfiles_clinicos_pacientes_silver_contest.xlsx")

    pacientes: dict[str, Paciente] = {}
    for fila in clinicos:
        pid = str(fila["paciente_id"])
        c = contacto.get(pid, {})
        pacientes[pid] = Paciente(
            paciente_id=pid,
            nombre_completo=str(c.get("nombre_completo") or "").strip(),
            documento_cc=str(c.get("documento_cc") or "").strip(),
            eps=str(c.get("eps") or "").strip(),
            ciudad=str(c.get("ciudad") or "").strip(),
            departamento=str(c.get("departamento") or "").strip(),
            direccion=str(c.get("direccion") or "").strip(),
            edad=int(fila.get("edad") or 0),
            genero=str(fila.get("genero") or "").strip(),
            procedimiento=str(fila.get("procedimiento") or "").strip(),
            modulo_synthea=str(fila.get("modulo_synthea") or "").strip(),
            fecha_cirugia=_como_fecha(fila.get("fecha_cirugia")),
            comorbilidades=_como_lista(fila.get("comorbilidades")),
        )
    return pacientes


def obtener_paciente(paciente_id: str) -> Paciente | None:
    return cargar_pacientes().get(paciente_id)


def listar_pacientes() -> list[Paciente]:
    return sorted(cargar_pacientes().values(), key=lambda p: p.paciente_id)


@dataclass(frozen=True)
class Ficha:
    """Ficha + día postoperatorio: el contexto exacto con el que se hace una llamada."""

    paciente: Paciente
    dia_postop: int

    def como_dict(self) -> dict[str, Any]:
        return {**self.paciente.como_dict(), "dia_postop": self.dia_postop}

    def resumen_para_prompt(self) -> str:
        """Bloque compacto que se inyecta en el prompt del agente."""
        p = self.paciente
        comorbilidades = ", ".join(p.comorbilidades) if p.comorbilidades else "ninguna registrada"
        return (
            f"Paciente: {p.nombre_completo} ({p.edad} años, {p.genero}).\n"
            f"Procedimiento: {p.procedimiento}.\n"
            f"Día postoperatorio: {self.dia_postop}.\n"
            f"Comorbilidades: {comorbilidades}.\n"
            f"EPS: {p.eps}. Ciudad: {p.ciudad}."
        )


def construir_ficha(paciente_id: str, dia_postop: int) -> Ficha | None:
    paciente = obtener_paciente(paciente_id)
    return Ficha(paciente=paciente, dia_postop=dia_postop) if paciente else None
